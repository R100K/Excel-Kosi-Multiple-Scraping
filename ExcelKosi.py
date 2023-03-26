#! python3
# ExcelKosi.py - Tool to automate boring tasks in Excel files
# Author - Robert Stokłosa

import openpyxl
from openpyxl.utils import get_column_letter

# Welcome instruction
print(f'KosiExcel program for scraping Excel files.'
      '\nPlace the Excel file to be checked in the folder where this program is located.\n')

# Opening Excel file
while True:
    fileName = input('Write the name of the file you want to check: ')
    try:
        file = openpyxl.load_workbook(fileName + '.xlsx')
        break
    except Exception:
        try:
            file = openpyxl.load_workbook(fileName + '.xltm')
            break
        except Exception:
            try:
                file = openpyxl.load_workbook(fileName + '.xlsm')
                break
            except Exception:
                try:
                    file = openpyxl.load_workbook(fileName + '.xltx')
                    break
                except Exception:
                    print('The file you provided does not exist or its name is incorrect'
                          '\nPlease try again')
                continue
            
# Choosing sheet
while True:
    try:
        fileSheet = input('Write the name of the sheet you want to search in: ')
        sheet = file[fileSheet]
    except Exception:
        print('The sheet you provided does not exist or its name is incorrect'
            '\nPlease try again')
        continue
    else:
        break

# Creating list of phrases to check
wordToCheck = input('Write the phrases that should be present in the file, separated by commas: ').split(", ")
print('')

# Creating text file to save results
f = open(fileName + 'ExcelScraping.txt', 'w', encoding="utf-8")
f. write(f'I am checking the file: %s \n' %(fileName))
f. write(f'Sheet: %s \n' %(fileSheet))
f. write(f'Looking for phrases: ' + ', '.join(wordToCheck) + '\n\n')
# Scraping file with list
for x in wordToCheck:
    i = 0
    for y in range(1, sheet.max_row+1):
        for z in range(1, sheet.max_column+1):
            if x.casefold() == str(sheet.cell(row=y, column=z).value).casefold():
                print(f"Phrase '%s' found in cell '%s%s'" %(sheet.cell(row=y, column=z).value, get_column_letter(z), y))
                f. write(f"Phrase '%s' found in cell '%s%s'\n" %(sheet.cell(row=y, column=z).value, get_column_letter(z), y))
                i += 1
    if i > 1:
        print(f"The phrase '%s' appears %s times in total\n" %(x, i))
        f. write(f"The phrase '%s' appears %s times in total\n\n" %(x, i))
    elif i == 1:
        print(f"The phrase '%s' appears %s time in total\n" %(x, i))
        f. write(f"The phrase '%s' appears %s time in total\n\n" %(x, i))
    else:
        print(f"The phrase '%s' does not appear in the file '%s'\n" %(x, fileName))
        f. write(f"The phrase '%s' does not appear in the file '%s'\n\n" %(x, fileName))
f. write(f"I wish you a pleasant work!")

# Saving results to txt file
f.close()
print(f"Report saved in the file '%sExcelScraping.txt'\n" %(fileName))
print(f"I wish you a pleasant work!\n")

# Closing program after pressing enter
input("Press ENTER to exit the program")